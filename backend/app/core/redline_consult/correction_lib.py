"""Shared helpers for the SOURCE_STATION_OCR_CORRECTION lane.

Dependency-light on purpose (only openpyxl/json/re, NO engine/coverage imports) so both
the standalone before/after tool (apply_and_rerun.py) and the coverage harness
(build_coverage.py) can use ONE implementation without circular imports.

Corrections are NON-MUTATING: the original .xlsx is never touched; a corrected COPY is
written and fed to the engine through the normal file-fed path. Only the explicit,
owner-approved per-cell station_map is applied — never a guessed value, never a blanket
digit swap.
"""
import json
import os
import re
from typing import Mapping

import openpyxl

_STA_PREFIX_RE = re.compile(r"(?i)^\s*sta\.?\s+")
_STA_FT_RE = re.compile(r"^(\d+)\+(\d+)$")


def _norm(s):
    return _STA_PREFIX_RE.sub("", str(s).strip()) if s is not None else s


def load_corrections(path):
    """Return {log_id: correction_dict} from a corrections.json (empty dict if missing)."""
    if not os.path.isfile(path):
        return {}
    data = json.load(open(path, encoding="utf-8"))
    return {c["log_id"]: c for c in data.get("corrections", [])}


def corrected_copy(src_xlsx, dst_dir, station_map, drop_stations=None, strict=True):
    """Write a corrected COPY of ``src_xlsx`` into ``dst_dir``, applying ONLY the explicit
    per-cell ``station_map`` to the 'station' column. The original is never modified.

    Returns ``(dst_path, [{row, original, corrected}])`` — every cell change recorded for
    audit. A cell is changed only when its (STA-stripped) text EXACTLY matches a station_map
    key; everything else is copied through verbatim.

    ``drop_stations`` (optional): exact source station strings whose ENTIRE row is excluded
    (superseded) from the corrected copy — for owner-reviewed mis-merged / contaminant rows
    that no value remap can fix. Matching is EXACT on the (STA-stripped) source value, never
    heuristic; a drop is keyed on the ORIGINAL source value and takes precedence over
    ``station_map`` for the same cell. Dropped rows are NOT listed in ``changes`` (they are
    not cell edits); the caller records them from its own owner-approved exclusion record.

    A correction must match the *current extracted value*: with ``strict=True`` (default), if
    any station_map key OR drop_station matches NO cell in the source, the correction is
    REFUSED with a ``ValueError`` (a stale/typo'd correction must never silently proceed, and
    no corrected copy is written). ``strict=False`` is for tests/dry-runs only.
    """
    os.makedirs(dst_dir, exist_ok=True)
    dst = os.path.join(dst_dir, os.path.basename(src_xlsx))
    drop_set = {_norm(s) for s in (drop_stations or [])}
    wb = openpyxl.load_workbook(src_xlsx)  # not data_only: edit + save real cell values
    try:
        ws = wb.worksheets[0]
        header = [(_norm(c.value).lower() if c.value is not None else "") for c in ws[1]]
        sc = header.index("station") if "station" in header else 0
        changes = []
        matched = set()
        matched_drops = set()
        rows_to_delete = []
        for row in ws.iter_rows(min_row=2):
            if sc >= len(row):
                continue
            cell = row[sc]
            key = _norm(cell.value)
            if key in drop_set:  # whole-row exclusion takes precedence over value remap
                matched_drops.add(key)
                rows_to_delete.append(cell.row)
                continue
            if key in station_map:
                matched.add(key)
                changes.append({"row": cell.row, "original": str(cell.value), "corrected": station_map[key]})
                cell.value = station_map[key]
        unmatched = sorted(set(station_map) - matched)
        if strict and unmatched:
            # refuse loudly; do NOT write a corrected copy from a correction that does not
            # match the current source (prevents silent / stale / typo'd corrections).
            raise ValueError(
                f"correction refused for {os.path.basename(src_xlsx)}: station_map keys not found "
                f"in the source (original value does not match current extracted value): {unmatched}")
        unmatched_drops = sorted(drop_set - matched_drops)
        if strict and unmatched_drops:
            raise ValueError(
                f"correction refused for {os.path.basename(src_xlsx)}: drop_stations not found "
                f"in the source (original value does not match current extracted value): {unmatched_drops}")
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r, 1)  # delete high->low so earlier row indices stay valid
        wb.save(dst)
    finally:
        wb.close()
    return dst, changes


def _station_ft(text):
    """``A+BB`` -> ``A*100+BB`` feet, else None (mirrors the engine station grammar; used only to
    keep the corrected COPY's ``station_ft`` consistent — the engine re-parses from ``station``)."""
    m = _STA_FT_RE.match(_norm(text) or "")
    return int(m.group(1)) * 100 + int(m.group(2)) if m else None


def corrected_rows(rows, station_map, drop_stations=None, strict=True):
    """Row-fed sibling of :func:`corrected_copy`: apply the SAME owner-approved per-cell
    ``station_map`` + whole-row ``drop_stations`` exclusions to TrueLine ``committed_rows`` (a
    list of dicts), returning a CORRECTED COPY. The input rows are NEVER mutated (each changed
    row is shallow-copied) — the live-app STATE committed_rows stay untouched, exactly as the
    file-fed path never alters the original .xlsx.

    Mirrors ``corrected_copy`` semantics exactly:
      * a cell changes only when its (STA-stripped) ``station`` text EXACTLY matches a
        ``station_map`` key (never a guess, never a blanket digit swap);
      * a row is dropped only when its ``station`` EXACTLY matches a ``drop_stations`` entry
        (a whole-row exclusion takes precedence over a value remap for the same cell);
      * with ``strict=True`` (default) any ``station_map`` key OR ``drop_stations`` entry that
        matches NO row REFUSES the correction with ``ValueError`` — a stale/typo'd correction
        can never silently proceed.

    Returns ``(new_rows, [{row, original, corrected}])`` — every cell change recorded for audit.
    """
    drop_set = {_norm(s) for s in (drop_stations or [])}
    out, changes = [], []
    matched, matched_drops = set(), set()
    for i, r in enumerate(rows or []):
        if not isinstance(r, Mapping):
            out.append(r)
            continue
        key = _norm(r.get("station"))
        if key in drop_set:  # whole-row exclusion takes precedence over value remap
            matched_drops.add(key)
            continue
        if key in station_map:
            matched.add(key)
            new_val = station_map[key]
            changes.append({"row": i, "original": str(r.get("station")), "corrected": new_val})
            nr = dict(r)
            nr["station"] = new_val
            if "station_ft" in nr:
                nr["station_ft"] = _station_ft(new_val)
            out.append(nr)
        else:
            out.append(r)
    unmatched = sorted(set(station_map) - matched)
    if strict and unmatched:
        raise ValueError(
            "correction refused: station_map keys not found in committed_rows "
            f"(original value does not match current extracted value): {unmatched}")
    unmatched_drops = sorted(drop_set - matched_drops)
    if strict and unmatched_drops:
        raise ValueError(
            "correction refused: drop_stations not found in committed_rows "
            f"(original value does not match current extracted value): {unmatched_drops}")
    return out, changes
