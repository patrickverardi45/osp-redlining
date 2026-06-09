"""ODOT VeroFy "Construction Log" bore-log reader -> canonical Bore.

The form is a key/value layout (label cell + value cell to its right) on a
``Construction Log`` sheet, carrying project, CD sheet(s), Start/End Station, and
Placed Footage. We read the clean authored fields only; station feet come from
the station text via the shared parser.
"""
from __future__ import annotations

import os
import re
from typing import Dict, List, Optional, Tuple

import openpyxl

from truelinev2.schema.models import Bore
from truelinev2.stations import parse_station


def _sheets(cd) -> List[int]:
    out: List[int] = []
    for n in re.findall(r"\d+", str(cd or "")):
        v = int(n)
        if v not in out:
            out.append(v)
    return out


def read_odot_borelog(path: str) -> Bore:
    wb = openpyxl.load_workbook(path, data_only=True)
    name = "Construction Log" if "Construction Log" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[name]
    rowvals: Dict[int, List[Tuple[int, str]]] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                rowvals.setdefault(cell.row, []).append((cell.column, str(cell.value).strip()))
    wb.close()

    def find(label: str) -> Optional[str]:
        for r in sorted(rowvals):
            cols = sorted(rowvals[r])
            for k, (_c, v) in enumerate(cols):
                if label.lower() in v.lower():
                    for _c2, v2 in cols[k + 1:]:
                        if v2 and label.lower() not in v2.lower():
                            return v2
        return None

    project = find("VeroFy Segment") or find("Construction Log #")
    cd = find("CD Sheets")
    start = find("Start Station")
    end = find("End Station")
    sf, ef = parse_station(start), parse_station(end)
    if sf is None or ef is None:
        raise ValueError(f"ODOT bore log missing parseable Start/End Station: {path}")
    lo, hi = (sf, ef) if sf <= ef else (ef, sf)
    lo_s, hi_s = (start, end) if sf <= ef else (end, start)
    base = os.path.basename(path)
    return Bore(
        bore_id=os.path.splitext(base)[0],
        project=project,
        source_file=base,
        sheet_refs=_sheets(cd),
        station_start=lo_s,
        station_end=hi_s,
        station_start_ft=lo,
        station_end_ft=hi,
        span_ft=round(hi - lo, 2),
        print_raw=cd,
    )
