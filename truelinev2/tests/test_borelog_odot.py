import openpyxl

from truelinev2.ingest.borelog_odot import read_odot_borelog
from truelinev2.ingest.normalize import detect_format


def _make(tmp_path):
    p = tmp_path / "TULSA 31 BORE LOG 118.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Instructions"
    wb.active["A1"] = "instructions..."
    ws = wb.create_sheet("Construction Log")

    def put(r, c, v):
        ws.cell(row=r, column=c, value=v)

    put(5, 1, "VeroFy Segment #"); put(5, 4, "ODOT TULSA 31")
    put(6, 1, "CD Sheets Names or"); put(6, 4, "Sheet 10 & 11")
    put(10, 1, "Start Station"); put(10, 4, "14+20"); put(10, 6, "End Station"); put(10, 9, "15+38")
    put(11, 1, "Placed Footage"); put(11, 4, "118'")
    wb.save(str(p))
    return str(p)


def test_detect_and_read_odot(tmp_path):
    p = _make(tmp_path)
    assert detect_format(p) == "odot_construction_log"
    bore = read_odot_borelog(p)
    assert bore.station_start_ft == 1420.0 and bore.station_end_ft == 1538.0
    assert bore.span_ft == 118.0
    assert bore.sheet_refs == [10, 11]
    assert bore.project == "ODOT TULSA 31"
